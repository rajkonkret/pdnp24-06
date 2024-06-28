import openpyxl

workbook = openpyxl.load_workbook('sales.xlsx')
worsheet = workbook.active
print(worsheet.title)  # Arkusz1

for i in worsheet:
    print(i)
# (<Cell 'Arkusz1'.A1>, <Cell 'Arkusz1'.B1>, <Cell 'Arkusz1'.C1>)
# (<Cell 'Arkusz1'.A2>, <Cell 'Arkusz1'.B2>, <Cell 'Arkusz1'.C2>)
# (<Cell 'Arkusz1'.A3>, <Cell 'Arkusz1'.B3>, <Cell 'Arkusz1'.C3>)
# (<Cell 'Arkusz1'.A4>, <Cell 'Arkusz1'.B4>, <Cell 'Arkusz1'.C4>)
# (<Cell 'Arkusz1'.A5>, <Cell 'Arkusz1'.B5>, <Cell 'Arkusz1'.C5>)

lista = []
for i in range(worsheet.max_row):
    for col in worsheet.iter_cols(1, worsheet.max_column):
        lista.append(col[i].value)

print(lista)
# ['Sales Date', 'Sales Person', 'Amount',
#  datetime.datetime(2018, 5, 12, 0, 0), 'Sila Ahmed', 60000,
#  datetime.datetime(2019, 12, 6, 0, 0), 'Mir Hossain', 50000,
#  datetime.datetime(2020, 8, 9, 0, 0), 'Sarmin Jahan', 45000,
#  datetime.datetime(2021, 4, 7, 0, 0), 'Mahmudul Hasan', 30000]
